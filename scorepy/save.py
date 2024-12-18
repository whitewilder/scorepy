import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill, Border, Side
from io import BytesIO

def save_charts(data_dict, file_name="output.xlsx"):
    """
    Save charts and tables into multiple sheets of an Excel file.

    Parameters:
        data_dict (dict): A dictionary where keys are sheet names, and values are tuples containing:
                          (list of charts, list of tables).
        file_name (str): Name of the output Excel file.
    """
    # Create a new Excel workbook
    wb = Workbook()
    first_sheet = True  # To check if it's the first sheet

    for sheet_name, (charts, tables) in data_dict.items():
        # Create a new sheet
        if first_sheet:
            ws = wb.active
            ws.title = sheet_name
            first_sheet = False
        else:
            ws = wb.create_sheet(title=sheet_name)
        
        current_row = 1  # Track the current row to place content sequentially

        # Add charts to the sheet
        for fig in charts:
            # Save the chart as an image in memory
            img_data = BytesIO()
            fig.savefig(img_data, format="png", bbox_inches="tight")
            img_data.seek(0)
            
            # Insert the chart image into the worksheet
            img = Image(img_data)
            img.anchor = f"A{current_row}"
            ws.add_image(img)
            
            # Leave space below the image
            current_row += 20
        
        # Add tables to the sheet
        for table in tables:
            # Round table values for better presentation
            table = table.round(2)

            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            border = Border(
                left=Side(style="thin"), 
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )

            # Write table header
            for col_idx, col_name in enumerate(table.columns, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border

            # Write table rows
            for row_idx, row in enumerate(table.itertuples(index=False), start=current_row + 1):
                for col_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
                    if isinstance(value, (int, float)):
                        cell.number_format = "0.00"

            # Leave space below the table
            current_row += len(table) + 3

    # Save the workbook
    wb.save(file_name)
    print(f"Saved multiple sheets to {file_name}")
